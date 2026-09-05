import os
import json
import logging
import datetime
import io
import subprocess
import shutil
import time
from fastapi import FastAPI, HTTPException, Header, Depends
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.request
import asyncio
import websockets
from bs4 import BeautifulSoup
import re

# ---------------------------------------------------------------------------
# DEBUG / LOGGING SETUP
# ---------------------------------------------------------------------------
# Set GMVN_DEBUG=0 in your environment to silence the verbose trail.
DEBUG_ENABLED = os.environ.get("GMVN_DEBUG", "1") != "0"

logger = logging.getLogger("gmvn_portal")
logger.setLevel(logging.DEBUG if DEBUG_ENABLED else logging.INFO)
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter(
        "[%(asctime)s] [%(levelname)s] %(message)s", datefmt="%H:%M:%S"
    ))
    logger.addHandler(_handler)


def dbg(msg: str):
    """Shorthand debug logger — every real-vs-fallback decision point logs through here."""
    if DEBUG_ENABLED:
        logger.debug(msg)


app = FastAPI(title="GMVN Real-Time Inventory & Tariff Matrix Portal", version="3.1.0-debug")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_FILE = os.path.join(os.path.dirname(__file__), "gmvn_master_accommodations.json")


def load_data():
    if not os.path.exists(DATA_FILE):
        dbg(f"DATA_FILE not found at {DATA_FILE} — starting with empty dataset.")
        return {"districts": [], "properties": []}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
        dbg(f"Loaded master dataset: {len(data.get('properties', []))} properties from {DATA_FILE}")
        return data


def generate_availability_for_dates(rooms, start_date_str, end_date_str, prop_seed=0):
    """
    FALLBACK / STATIC generator. This NEVER touches the live website.
    Every value here comes from gmvn_master_accommodations.json (or the
    in-memory CACHED_DATA, which /api/live-sync may have previously updated).
    """
    try:
        start_d = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_d = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        if end_d < start_d:
            end_d = start_d
    except Exception:
        start_d = datetime.date.today()
        end_d = start_d + datetime.timedelta(days=1)

    date_list = []
    delta_days = (end_d - start_d).days + 1
    if delta_days > 30:
        delta_days = 30
    for i in range(delta_days):
        d_val = start_d + datetime.timedelta(days=i)
        date_list.append(d_val.strftime("%Y-%m-%d"))

    processed_rooms = []
    for r_idx, r in enumerate(rooms):
        daily_inv = {}
        base_cap = r.get("base_available", r.get("available", 0))

        for d_str in date_list:
            if "daily_inventory" in r and d_str in r["daily_inventory"]:
                daily_inv[d_str] = r["daily_inventory"][d_str]
            else:
                daily_inv[d_str] = base_cap

        processed_rooms.append({
            "name": r["name"],
            "code": r.get("code", ""),
            "tariff": r["tariff"],
            "plan": r.get("plan", "EP Plan"),
            "contact": r.get("contact", "0135-2430373"),
            "base_available": base_cap,
            "daily_inventory": daily_inv,
            # --- DEBUG FIELDS ---
            "data_source": "FALLBACK_STATIC",
            "debug_reason": "Served from gmvn_master_accommodations.json / in-memory cache (no live scrape attempted or scrape failed)."
        })

    dbg(f"generate_availability_for_dates(): built {len(processed_rooms)} FALLBACK_STATIC room rows "
        f"for seed={prop_seed}, dates={date_list[0] if date_list else 'NA'}..{date_list[-1] if date_list else 'NA'}")

    return date_list, processed_rooms


# Master in-memory data cache
CACHED_DATA = load_data()

# ---------------------------------------------------------------------------
# LIVE_SCRAPE_CACHE persistence + TTL
# ---------------------------------------------------------------------------
# Cache entries are stored as: {cache_key: {"rooms": [...], "scraped_at": <iso8601>}}
# so every consumer knows exactly how old a "live" result actually is, and
# stale entries can be treated as a miss instead of served forever.
LIVE_CACHE_FILE = os.path.join(os.path.dirname(__file__), "live_scrape_cache.json")
LIVE_CACHE_TTL_SECONDS = int(os.environ.get("GMVN_CACHE_TTL_SECONDS", "3600"))  # 1 hour default


def _load_live_cache_from_disk():
    if not os.path.exists(LIVE_CACHE_FILE):
        return {}
    try:
        with open(LIVE_CACHE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            dbg(f"Loaded {len(data)} entries from persisted LIVE_SCRAPE_CACHE at {LIVE_CACHE_FILE}")
            return data
    except Exception as e:
        dbg(f"Failed to load persisted LIVE_SCRAPE_CACHE: {e}")
        return {}


def _save_live_cache_to_disk():
    try:
        with open(LIVE_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(LIVE_SCRAPE_CACHE, f, indent=2)
    except Exception as e:
        dbg(f"Failed to persist LIVE_SCRAPE_CACHE to disk: {e}")


LIVE_SCRAPE_CACHE = _load_live_cache_from_disk()


def _cache_entry_age_seconds(entry) -> float:
    try:
        scraped_at = datetime.datetime.fromisoformat(entry["scraped_at"])
        return (datetime.datetime.utcnow() - scraped_at).total_seconds()
    except Exception:
        return float("inf")  # malformed/unknown age -> treat as expired


def _cache_get_fresh(cache_key: str):
    """Returns the rooms list for cache_key only if present AND not expired; else None."""
    entry = LIVE_SCRAPE_CACHE.get(cache_key)
    if not entry:
        return None
    age = _cache_entry_age_seconds(entry)
    if age > LIVE_CACHE_TTL_SECONDS:
        dbg(f"Cache entry for key='{cache_key}' is stale (age={age:.0f}s > TTL={LIVE_CACHE_TTL_SECONDS}s) — treating as miss.")
        return None
    return entry["rooms"]


def _cache_set(cache_key: str, rooms):
    LIVE_SCRAPE_CACHE[cache_key] = {
        "rooms": rooms,
        "scraped_at": datetime.datetime.utcnow().isoformat()
    }
    _save_live_cache_to_disk()


# ---------------------------------------------------------------------------
# Simple API-key protection for state-changing endpoints
# ---------------------------------------------------------------------------
# Set GMVN_API_KEY in your environment to require a matching X-API-Key header
# on /api/live-sync, /api/live-sync/retry-failed, and /api/capture-session-cookies.
# If GMVN_API_KEY is unset, these endpoints remain open (dev-friendly default).
API_KEY = os.environ.get("GMVN_API_KEY", "").strip()


def require_api_key(x_api_key: Optional[str] = Header(default=None)):
    if API_KEY and x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="Missing or invalid X-API-Key header.")
    return True


# ---------------------------------------------------------------------------
# Chrome CDP self-healing: try to relaunch Chrome with remote debugging if
# the bridge is unreachable, instead of silently falling back every time.
# ---------------------------------------------------------------------------
CHROME_DEBUG_PORT = 9222
CHROME_PATH = os.environ.get("GMVN_CHROME_PATH", "").strip() or None
_DEFAULT_CHROME_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    "/usr/bin/google-chrome",
    "/usr/bin/chromium-browser",
]


def _find_chrome_binary() -> Optional[str]:
    if CHROME_PATH and os.path.exists(CHROME_PATH):
        return CHROME_PATH
    for candidate in _DEFAULT_CHROME_CANDIDATES:
        if os.path.exists(candidate):
            return candidate
    found = shutil.which("chrome") or shutil.which("google-chrome") or shutil.which("chromium-browser")
    return found


def _cdp_is_reachable(timeout=1.5) -> bool:
    try:
        req = urllib.request.Request(f"http://127.0.0.1:{CHROME_DEBUG_PORT}/json", headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=timeout) as r:
            json.loads(r.read())
        return True
    except Exception:
        return False


def try_relaunch_chrome_debug() -> dict:
    """
    Best-effort attempt to relaunch Chrome with the remote-debugging port open
    when the CDP bridge is unreachable. Returns a dict describing what happened
    so callers can log/report it rather than silently failing.
    """
    if _cdp_is_reachable():
        return {"attempted": False, "reason": "CDP already reachable, no relaunch needed."}

    chrome_bin = _find_chrome_binary()
    if not chrome_bin:
        return {
            "attempted": False,
            "reason": "No Chrome binary found (set GMVN_CHROME_PATH env var to its full path)."
        }

    try:
        subprocess.Popen(
            [chrome_bin, f"--remote-debugging-port={CHROME_DEBUG_PORT}", "--no-first-run", "--new-window", "about:blank"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )
        dbg(f"Attempted to relaunch Chrome with CDP debug port from: {chrome_bin}")
    except Exception as e:
        return {"attempted": True, "reason": f"Launch failed: {e}", "chrome_path": chrome_bin, "reachable_after": False}

    # Give Chrome a few seconds to start up before checking again.
    for _ in range(6):
        time.sleep(1)
        if _cdp_is_reachable():
            return {"attempted": True, "reason": "Relaunched successfully.", "chrome_path": chrome_bin, "reachable_after": True}

    return {"attempted": True, "reason": "Launched but CDP still unreachable after 6s.", "chrome_path": chrome_bin, "reachable_after": False}


# ---------------------------------------------------------------------------
# Last live-sync failure tracking (powers /api/live-sync/retry-failed)
# ---------------------------------------------------------------------------
LAST_SYNC_FAILED_TRH_IDS = []
LAST_SYNC_PARAMS = {}


# Global Session Cookie Storage & Scraping Config
ACTIVE_SESSION_COOKIES = []
DEFAULT_INTER_REQUEST_DELAY_MS = 1200
DEFAULT_CONCURRENCY_LIMIT = 2
DEFAULT_RATE_LIMIT_RETRIES = 3

@app.post("/api/capture-session-cookies")
async def capture_session_cookies(_auth: bool = Depends(require_api_key)):
    """
    Connects to Chrome CDP, extracts active cf_clearance, PHPSESSID, and security tokens from GMVN.
    """
    global ACTIVE_SESSION_COOKIES
    try:
        if not _cdp_is_reachable():
            relaunch_result = try_relaunch_chrome_debug()
            dbg(f"CDP unreachable before capturing cookies — relaunch attempt result: {relaunch_result}")

        req = urllib.request.Request("http://127.0.0.1:9222/json", headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=1.5) as r:
            tabs = json.loads(r.read())
        if not tabs:
            raise Exception("No active Chrome tabs found on 127.0.0.1:9222")
        ws_url = tabs[0]["webSocketDebuggerUrl"]

        async with websockets.connect(ws_url, max_size=10 * 1024 * 1024) as ws:
            await ws.send(json.dumps({"id": 1, "method": "Network.getCookies", "params": {"urls": ["https://gmvnonline.com"]}}))
            resp = json.loads(await ws.recv())
            cookies = resp.get("result", {}).get("cookies", [])
            ACTIVE_SESSION_COOKIES = cookies
            dbg(f"Captured {len(cookies)} live session cookies from Chrome CDP.")
            
            return {
                "success": True,
                "cookie_count": len(cookies),
                "cookies": [{"name": c["name"], "domain": c["domain"], "expires": c.get("expires")} for c in cookies],
                "message": "Live session cookies captured successfully."
            }
    except Exception as e:
        dbg(f"Failed to capture session cookies via CDP: {e}")
        return {
            "success": False,
            "cookie_count": len(ACTIVE_SESSION_COOKIES),
            "error": str(e),
            "message": "Could not connect to Chrome CDP on port 9222."
        }

def default_checkin_checkout():
    """
    Computes sensible default check-in/check-out dates relative to "today"
    (a week out, 5-night stay) instead of a hardcoded date that silently
    goes stale (and, once in the past, breaks live scraping against the
    GMVN portal, which returns 0 rooms for past-dated searches).
    """
    today = datetime.date.today()
    cin = today + datetime.timedelta(days=7)
    cout = cin + datetime.timedelta(days=5)
    return cin.strftime("%Y-%m-%d"), cout.strftime("%Y-%m-%d")


def normalize_room_name(name: str) -> str:
    """
    Normalizes a room name for fuzzy matching between the master dataset's
    room list and room names scraped from the live GMVN portal (which may
    differ in casing, punctuation, or extra whitespace).
    """
    if not name:
        return ""
    n = name.lower().strip()
    n = re.sub(r"[^a-z0-9\s]", "", n)   # strip punctuation
    n = re.sub(r"\s+", " ", n).strip()  # collapse whitespace
    return n


def get_cached_live_rooms(trh_id: str, checkin_str: str, checkout_str: str):
    """Cache-only lookup — never triggers a network scrape. Returns rooms list or None (also None if the entry has expired past the TTL)."""
    try:
        cin_d = datetime.datetime.strptime(checkin_str, "%Y-%m-%d").strftime("%d-%m-%Y")
        cout_d = datetime.datetime.strptime(checkout_str, "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        cin_d = checkin_str
        cout_d = checkout_str
    cache_key = f"{trh_id}_{cin_d}_{cout_d}"
    return _cache_get_fresh(cache_key)


async def scrape_gmvn_live_room_tariff(
    trh_id: str,
    checkin_str: str,
    checkout_str: str,
    inter_delay_ms: int = DEFAULT_INTER_REQUEST_DELAY_MS,
    retries: int = DEFAULT_RATE_LIMIT_RETRIES
):
    """
    Connects to Chrome CDP (ws://127.0.0.1:9222).
    Extracts dynamic security tokens, navigates with human-like delay, handles Turnstile challenge, and parses live room cards.
    """
    try:
        cin_d = datetime.datetime.strptime(checkin_str, "%Y-%m-%d").strftime("%d-%m-%Y")
        cout_d = datetime.datetime.strptime(checkout_str, "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        cin_d = checkin_str
        cout_d = checkout_str

    cache_key = f"{trh_id}_{cin_d}_{cout_d}"
    _fresh = _cache_get_fresh(cache_key)
    if _fresh is not None:
        dbg(f"[TRH {trh_id}] Cache HIT for key '{cache_key}' — returning cached live data.")
        return _fresh, "CACHED_LIVE", f"Served from LIVE_SCRAPE_CACHE, key={cache_key}"

    # Human-like inter-request pacing delay
    if inter_delay_ms > 0:
        await asyncio.sleep(inter_delay_ms / 1000.0)

    for attempt in range(retries):
        try:
            if not _cdp_is_reachable():
                relaunch_result = try_relaunch_chrome_debug()
                dbg(f"[TRH {trh_id}] CDP unreachable before attempt {attempt+1} — relaunch attempt result: {relaunch_result}")
                if not relaunch_result.get("reachable_after") and not _cdp_is_reachable():
                    return None, "FALLBACK_STATIC", f"CDP endpoint unreachable at 127.0.0.1:9222 and relaunch attempt did not restore it ({relaunch_result.get('reason')})."

            req = urllib.request.Request("http://127.0.0.1:9222/json", headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=1.5) as r:
                tabs = json.loads(r.read())
            if not tabs:
                return None, "FALLBACK_STATIC", "CDP endpoint returned no open tabs at 127.0.0.1:9222."
            ws_url = tabs[0]["webSocketDebuggerUrl"]

            async with websockets.connect(ws_url, max_size=10 * 1024 * 1024) as ws:
                # 1. Fetch current dynamic token from active page if present
                await ws.send(json.dumps({
                    "id": 1,
                    "method": "Runtime.evaluate",
                    "params": {"expression": "(document.querySelector('input[name=\"log\"]') || {}).value || '23f69e30bb19218a'", "returnByValue": True}
                }))
                token_resp = json.loads(await ws.recv())
                log_token = token_resp.get("result", {}).get("result", {}).get("value", "23f69e30bb19218a")

                target_url = f"https://gmvnonline.com/room-tariff.php?trhID={trh_id}&checkindate={cin_d}&checkoutdate={cout_d}&adults=&child=&log={log_token}"
                dbg(f"[TRH {trh_id}] Navigating via CDP (Attempt {attempt+1}/{retries}): {target_url}")
                
                await ws.send(json.dumps({"id": 2, "method": "Page.navigate", "params": {"url": target_url}}))
                await ws.recv()

                # Dynamic polling to resolve Cloudflare Turnstile challenge
                for _ in range(12):
                    await asyncio.sleep(0.8)
                    await ws.send(json.dumps({"id": 100, "method": "Runtime.evaluate", "params": {"expression": "document.title", "returnByValue": True}}))
                    t_resp = json.loads(await ws.recv())
                    cur_title = t_resp.get("result", {}).get("result", {}).get("value", "")
                    if cur_title and "Just a moment" not in cur_title:
                        break

                await ws.send(json.dumps({
                    "id": 3,
                    "method": "Runtime.evaluate",
                    "params": {"expression": "document.documentElement.outerHTML", "returnByValue": True}
                }))
                resp = json.loads(await ws.recv())
                html = resp.get("result", {}).get("result", {}).get("value", "")

                soup = BeautifulSoup(html, "html.parser")
                rooms = []
                seen = set()
                boxes = soup.find_all("div", class_="blog-content")

                for box in boxes:
                    h3 = box.find("h3")
                    if not h3 or "tariff" in h3.text.lower():
                        continue
                    rname = h3.text.strip()
                    if rname in seen:
                        continue
                    seen.add(rname)

                    parent = box.find_parent("div", class_="row") or box.parent
                    tariff = 0
                    avail = 0
                    contact = "9568006683"
                    plan = "CP Plan (Bed Tea & Breakfast)"
                    if parent:
                        text = parent.text
                        for line in text.splitlines():
                            l = line.strip()
                            if not l: continue
                            if "tariff" in l.lower() or "₹" in l:
                                nums = re.findall(r'\d+', l.replace(",", ""))
                                if nums and tariff == 0: tariff = int(nums[0])
                            if "available room" in l.lower() or "available bed" in l.lower():
                                nums = re.findall(r'\d+', l)
                                if nums: avail = int(nums[0])
                            if "not available" in l.lower():
                                avail = 0
                            if "contact no" in l.lower():
                                m_nums = re.findall(r'[\d\s,]+', l)
                                if m_nums: contact = m_nums[0].strip()
                            if "plan" in l.lower():
                                plan = l

                    rooms.append({
                        "name": rname,
                        "tariff": tariff,
                        "available": avail,
                        "contact": contact,
                        "plan": plan
                    })

                if rooms:
                    _cache_set(cache_key, rooms)
                    dbg(f"[TRH {trh_id}] SUCCESS — scraped {len(rooms)} real room rows from live portal.")
                    return rooms, "LIVE_SCRAPE", f"Successfully parsed {len(rooms)} rooms from live HTML."

        except Exception as e:
            dbg(f"[TRH {trh_id}] Attempt {attempt+1} encountered error: {e}")
            await asyncio.sleep(1.0)

    return None, "FALLBACK_STATIC", "HTML fetched but 0 rooms matched or request timed out."


@app.get("/api/inventory-matrix")
async def get_inventory_matrix(
    district: Optional[str] = None,
    city_search: Optional[str] = None,
    checkin: Optional[str] = None,
    checkout: Optional[str] = None
):
    global CACHED_DATA
    if not checkin or not checkout:
        _def_cin, _def_cout = default_checkin_checkout()
        checkin = checkin or _def_cin
        checkout = checkout or _def_cout
    props = CACHED_DATA.get("properties", [])
    dbg(f"=== /api/inventory-matrix called === district={district!r}, city_search={city_search!r}, "
        f"checkin={checkin}, checkout={checkout}")

    if district and district.lower() != "all":
        props = [p for p in props if p["district"].lower() == district.lower()]

    if city_search:
        s = city_search.lower().strip()
        props = [p for p in props if s in p["name"].lower() or s in p["city"].lower() or s in p["district"].lower()]

    dbg(f"Filtered property count: {len(props)}")

    try:
        start_d = datetime.datetime.strptime(checkin, "%Y-%m-%d").date()
        end_d = datetime.datetime.strptime(checkout, "%Y-%m-%d").date()
        if end_d < start_d:
            end_d = start_d
    except Exception:
        start_d = datetime.date.today()
        end_d = start_d + datetime.timedelta(days=1)

    date_list = []
    delta_days = min(31, (end_d - start_d).days + 1)
    for i in range(delta_days):
        date_list.append((start_d + datetime.timedelta(days=i)).strftime("%Y-%m-%d"))
    all_dates = date_list

    results = []
    # Live scraping is only attempted for narrow, specific searches.
    do_live_scrape = bool(city_search and len(props) <= 3)
    dbg(f"do_live_scrape={do_live_scrape} "
        f"(requires: city_search set AND filtered property count <= 3). "
        f"Current: city_search={'set' if city_search else 'NOT set'}, prop_count={len(props)}")

    stats = {"live_scrape": 0, "cached_live": 0, "fallback_static": 0}

    for p in props:
        trh_id = p.get("trh_id", "")
        live_rooms = None
        source_tag = "FALLBACK_STATIC"
        reason = "do_live_scrape=False for this request (broad search / district-only browse)."

        if do_live_scrape:
            live_rooms, source_tag, reason = await scrape_gmvn_live_room_tariff(trh_id, checkin, checkout)
        else:
            # Even when we won't trigger a fresh live scrape (broad browse),
            # still serve already-cached data from a prior /api/live-sync run
            # instead of always falling back to static placeholder rows.
            cached_rooms = get_cached_live_rooms(trh_id, checkin, checkout)
            if cached_rooms:
                live_rooms = cached_rooms
                source_tag = "CACHED_LIVE"
                reason = "Served from LIVE_SCRAPE_CACHE (populated by a prior /api/live-sync run); do_live_scrape=False so no fresh scrape was attempted."

        dbg(f"[TRH {trh_id}] property='{p.get('name')}' -> source={source_tag} | reason: {reason}")

        rooms_output = []
        if live_rooms:
            for lr in live_rooms:
                # NOTE (debug flag): this stamps ONE scraped snapshot value across
                # every date in the requested range — it is not a true per-day
                # calendar. Flagged here so it's visible in the API output.
                daily_inv = {d_str: lr["available"] for d_str in date_list}
                rooms_output.append({
                    "name": lr["name"],
                    "code": lr["name"][:5].upper().replace(" ", "-"),
                    "tariff": lr["tariff"],
                    "plan": lr["plan"],
                    "contact": lr["contact"],
                    "base_available": lr["available"],
                    "daily_inventory": daily_inv,
                    "data_source": source_tag,
                    "debug_reason": reason + " [WARNING: single snapshot value repeated across all requested dates, not true daily granularity]"
                })
            if source_tag == "LIVE_SCRAPE":
                stats["live_scrape"] += 1
            elif source_tag == "CACHED_LIVE":
                stats["cached_live"] += 1
        else:
            _, rooms_output = generate_availability_for_dates(
                p.get("rooms", []),
                checkin or "2026-09-25",
                checkout or "2026-09-30",
                prop_seed=int(p.get("trh_id", 1))
            )
            # Attach the specific reason this property fell back (overwrite generic default)
            for r in rooms_output:
                r["debug_reason"] = reason
            stats["fallback_static"] += 1

        results.append({
            "trh_id": p["trh_id"],
            "district": p["district"],
            "city": p["city"],
            "property_name": p["name"],
            "rooms": rooms_output
        })

    dbg(f"=== Request summary: {stats['live_scrape']} live-scraped, {stats['cached_live']} cached-live, "
        f"{stats['fallback_static']} fallback-static (out of {len(props)} properties) ===")

    return {
        "dates": all_dates,
        "total_properties": len(results),
        "properties": results,
        "_debug_stats": {
            "debug_enabled": DEBUG_ENABLED,
            "do_live_scrape_attempted": do_live_scrape,
            "live_scrape_count": stats["live_scrape"],
            "cached_live_count": stats["cached_live"],
            "fallback_static_count": stats["fallback_static"],
            "date_granularity_warning": "For LIVE_SCRAPE / CACHED_LIVE rooms, a single scraped availability/tariff "
                "snapshot (for the requested checkin/checkout pair) is repeated across every date in the response's "
                "'dates' array. This is NOT true per-day granularity — GMVN's portal returns one snapshot per "
                "trhID+date-range request, not a day-by-day calendar. FALLBACK_STATIC rooms use whatever "
                "daily_inventory values exist in the master dataset, which may or may not vary by day.",
            "cache_ttl_seconds": LIVE_CACHE_TTL_SECONDS,
            "note": "do_live_scrape is only True when city_search is set AND the filtered result set is <= 3 properties. "
                    "Otherwise every property below is FALLBACK_STATIC unless already present (and unexpired) in LIVE_SCRAPE_CACHE."
        }
    }


async def _run_live_sync_for_properties(props, checkin, checkout, inter_delay_ms, retries, concurrency_limit):
    """
    Shared scraping-loop core used by both /api/live-sync (full batch) and
    /api/live-sync/retry-failed (just the properties that failed last time).
    Returns (scraped_count, failed_count, failed_trh_ids, per_property_log).
    """
    semaphore = asyncio.Semaphore(max(1, min(10, concurrency_limit or 2)))
    scraped_count = 0
    failed_count = 0
    failed_trh_ids = []
    per_property_log = []

    async def scrape_property_worker(p):
        nonlocal scraped_count, failed_count
        trh_id = p.get("trh_id", "")
        async with semaphore:
            rooms, source_tag, reason = await scrape_gmvn_live_room_tariff(
                trh_id, checkin, checkout, inter_delay_ms=inter_delay_ms or 1200, retries=retries or 3
            )
            per_property_log.append({
                "trh_id": trh_id, "property_name": p.get("name"), "source": source_tag, "reason": reason
            })
            if rooms:
                scraped_count += 1
                for r in p.get("rooms", []):
                    norm_r = normalize_room_name(r["name"])
                    for lr in rooms:
                        norm_lr = normalize_room_name(lr["name"])
                        if norm_r == norm_lr or norm_r in norm_lr or norm_lr in norm_r:
                            r["base_available"] = lr["available"]
                            r["tariff"] = lr["tariff"]
            else:
                failed_count += 1
                failed_trh_ids.append(trh_id)

    tasks = [scrape_property_worker(p) for p in props]
    if tasks:
        await asyncio.gather(*tasks)

    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(CACHED_DATA, f, indent=2)
    except Exception as e:
        dbg(f"Failed to persist live-synced master data: {e}")

    return scraped_count, failed_count, failed_trh_ids, per_property_log


@app.post("/api/live-sync")
async def live_sync(
    district: Optional[str] = None,
    city_search: Optional[str] = None,
    checkin: Optional[str] = None,
    checkout: Optional[str] = None,
    inter_delay_ms: Optional[int] = DEFAULT_INTER_REQUEST_DELAY_MS,
    retries: Optional[int] = DEFAULT_RATE_LIMIT_RETRIES,
    concurrency_limit: Optional[int] = DEFAULT_CONCURRENCY_LIMIT,
    _auth: bool = Depends(require_api_key)
):
    """
    Clears live cache, captures fresh session tokens, scrapes filtered GMVN properties with concurrency & pacing limits, and updates in-memory feed.
    """
    global CACHED_DATA, LIVE_SCRAPE_CACHE, LAST_SYNC_FAILED_TRH_IDS, LAST_SYNC_PARAMS
    if not checkin or not checkout:
        _def_cin, _def_cout = default_checkin_checkout()
        checkin = checkin or _def_cin
        checkout = checkout or _def_cout
    dbg(f"=== /api/live-sync called === district={district!r}, city_search={city_search!r}, "
        f"checkin={checkin}, checkout={checkout}, inter_delay_ms={inter_delay_ms}, concurrency={concurrency_limit}")

    LIVE_SCRAPE_CACHE.clear()
    _save_live_cache_to_disk()

    # Pre-flight: Capture live session cookies
    await capture_session_cookies()

    props = CACHED_DATA.get("properties", [])
    if district and district.lower() != "all":
        props = [p for p in props if p["district"].lower() == district.lower()]
    if city_search:
        s = city_search.lower().strip()
        props = [p for p in props if s in p["name"].lower() or s in p["city"].lower() or s in p["district"].lower()]

    dbg(f"live_sync will scrape {len(props)} properties with concurrency={concurrency_limit} and delay={inter_delay_ms}ms")

    scraped_count, failed_count, failed_trh_ids, per_property_log = await _run_live_sync_for_properties(
        props, checkin, checkout, inter_delay_ms, retries, concurrency_limit
    )

    LAST_SYNC_FAILED_TRH_IDS = failed_trh_ids
    LAST_SYNC_PARAMS = {
        "district": district, "city_search": city_search, "checkin": checkin, "checkout": checkout,
        "inter_delay_ms": inter_delay_ms, "retries": retries, "concurrency_limit": concurrency_limit
    }

    dbg(f"=== live-sync summary: {scraped_count} succeeded, {failed_count} failed (out of {len(props)}) ===")

    return {
        "success": True,
        "scraped_properties_count": scraped_count,
        "failed_properties_count": failed_count,
        "failed_trh_ids": failed_trh_ids,
        "checkin": checkin,
        "checkout": checkout,
        "cookies_captured": len(ACTIVE_SESSION_COOKIES),
        "inter_delay_ms": inter_delay_ms,
        "concurrency_limit": concurrency_limit,
        "_debug_per_property": per_property_log
    }


@app.post("/api/live-sync/retry-failed")
async def live_sync_retry_failed(_auth: bool = Depends(require_api_key)):
    """
    Re-scrapes only the properties that failed on the last /api/live-sync run,
    reusing the same checkin/checkout/pacing parameters — instead of forcing a
    full re-scrape of all properties (including ones that already succeeded)
    just to pick up the handful that timed out.
    """
    global LAST_SYNC_FAILED_TRH_IDS, LAST_SYNC_PARAMS
    if not LAST_SYNC_FAILED_TRH_IDS:
        return {
            "success": True,
            "message": "No failed properties recorded from the last /api/live-sync run — nothing to retry.",
            "scraped_properties_count": 0,
            "failed_properties_count": 0
        }
    if not LAST_SYNC_PARAMS:
        raise HTTPException(status_code=400, detail="No prior /api/live-sync run found in this server session.")

    checkin = LAST_SYNC_PARAMS.get("checkin")
    checkout = LAST_SYNC_PARAMS.get("checkout")
    inter_delay_ms = LAST_SYNC_PARAMS.get("inter_delay_ms")
    retries = LAST_SYNC_PARAMS.get("retries")
    concurrency_limit = LAST_SYNC_PARAMS.get("concurrency_limit")

    all_props = CACHED_DATA.get("properties", [])
    retry_props = [p for p in all_props if p.get("trh_id") in LAST_SYNC_FAILED_TRH_IDS]

    dbg(f"=== /api/live-sync/retry-failed called === retrying {len(retry_props)} properties: {LAST_SYNC_FAILED_TRH_IDS}")

    await capture_session_cookies()

    scraped_count, failed_count, failed_trh_ids, per_property_log = await _run_live_sync_for_properties(
        retry_props, checkin, checkout, inter_delay_ms, retries, concurrency_limit
    )

    LAST_SYNC_FAILED_TRH_IDS = failed_trh_ids

    dbg(f"=== retry-failed summary: {scraped_count} succeeded, {failed_count} still failed (out of {len(retry_props)}) ===")

    return {
        "success": True,
        "retried_count": len(retry_props),
        "scraped_properties_count": scraped_count,
        "failed_properties_count": failed_count,
        "failed_trh_ids": failed_trh_ids,
        "checkin": checkin,
        "checkout": checkout,
        "_debug_per_property": per_property_log
    }


@app.get("/api/export/excel")
async def export_excel(
    district: Optional[str] = None,
    city_search: Optional[str] = None,
    checkin: Optional[str] = None,
    checkout: Optional[str] = None
):
    if not checkin or not checkout:
        _def_cin, _def_cout = default_checkin_checkout()
        checkin = checkin or _def_cin
        checkout = checkout or _def_cout
    matrix_data = await get_inventory_matrix(district, city_search, checkin, checkout)
    dates = matrix_data.get("dates", [])
    props = matrix_data.get("properties", [])
    dbg(f"export_excel(): building workbook for {len(props)} properties, "
        f"debug_stats={matrix_data.get('_debug_stats')}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GMVN Inventory & Availability"

    header_fill = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    date_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    date_font = Font(name="Segoe UI", size=10, bold=True, color="F59E0B")

    prop_font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    room_font = Font(name="Segoe UI", size=10, color="1E293B")
    avail_font = Font(name="Segoe UI", size=10, bold=True, color="059669")
    sold_font = Font(name="Segoe UI", size=10, bold=True, color="DC2626")
    source_live_font = Font(name="Segoe UI", size=9, bold=True, color="047857")
    source_fallback_font = Font(name="Segoe UI", size=9, bold=True, color="92400E")

    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # NOTE: "Data Source" column added (position 8) so any exported sheet is
    # self-auditing — you can see per-row whether it was LIVE_SCRAPE,
    # CACHED_LIVE, or FALLBACK_STATIC without needing the server logs.
    headers = ["District", "City / Area", "Property Name", "Room Category",
               "Tariff (INR)", "Meal Plan", "Contact No.", "Data Source"]
    for d in dates:
        d_obj = datetime.datetime.strptime(d, "%Y-%m-%d")
        headers.append(d_obj.strftime("%d-%b (%a)"))

    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        if col_idx <= 8:
            cell.fill = header_fill
            cell.font = header_font
        else:
            cell.fill = date_fill
            cell.font = date_font

    ws.row_dimensions[1].height = 28

    current_row = 2
    for p in props:
        for r in p["rooms"]:
            source_tag = r.get("data_source", "FALLBACK_STATIC")
            source_label = {
                "LIVE_SCRAPE": "LIVE (scraped now)",
                "CACHED_LIVE": "LIVE (cached)",
                "FALLBACK_STATIC": "STATIC fallback"
            }.get(source_tag, source_tag)

            row_data = [
                p["district"],
                p["city"],
                p["property_name"],
                r["name"],
                f"₹ {r['tariff']:,}",
                r["plan"],
                r["contact"],
                source_label
            ]
            for d in dates:
                avail_count = r["daily_inventory"].get(d, 0)
                row_data.append(f"{avail_count} Rooms" if avail_count > 0 else "Sold Out")

            ws.append(row_data)

            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border_thin
                cell.alignment = Alignment(vertical="center")

                if col_idx in [1, 2, 3]:
                    cell.font = prop_font
                elif col_idx in [4, 5, 6, 7]:
                    cell.font = room_font
                    if col_idx == 5:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 8:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if source_tag == "FALLBACK_STATIC":
                        cell.font = source_fallback_font
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    else:
                        cell.font = source_live_font
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if "Sold Out" in str(cell.value):
                        cell.font = sold_font
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    else:
                        cell.font = avail_font
                        cell.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

            ws.row_dimensions[current_row].height = 20
            current_row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"GMVN_Inventory_Availability_{checkin}_to_{checkout}.xlsx"
    dbg(f"export_excel(): workbook built, {current_row - 2} data rows written, file='{filename}'")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/debug/status")
async def debug_status():
    """
    Quick health-check endpoint: tells you right now, without touching any
    property, whether the CDP-controlled Chrome bridge is even reachable —
    the single biggest reason live scraping silently falls back.
    """
    cdp_reachable = _cdp_is_reachable()
    cdp_error = None if cdp_reachable else "CDP endpoint not reachable at 127.0.0.1:9222"
    tab_count = 0
    if cdp_reachable:
        try:
            req = urllib.request.Request("http://127.0.0.1:9222/json", headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=1.5) as r:
                tab_count = len(json.loads(r.read()))
        except Exception as e:
            cdp_error = f"{type(e).__name__}: {e}"
            cdp_reachable = False

    cache_details = []
    now_ = datetime.datetime.utcnow()
    for k, v in LIVE_SCRAPE_CACHE.items():
        age = _cache_entry_age_seconds(v)
        cache_details.append({
            "key": k,
            "age_seconds": round(age, 1) if age != float("inf") else None,
            "expired": age > LIVE_CACHE_TTL_SECONDS,
            "scraped_at": v.get("scraped_at")
        })

    status = {
        "debug_enabled": DEBUG_ENABLED,
        "cdp_bridge_reachable": cdp_reachable,
        "cdp_open_tab_count": tab_count,
        "cdp_error": cdp_error,
        "chrome_auto_relaunch_available": _find_chrome_binary() is not None,
        "chrome_binary_detected": _find_chrome_binary(),
        "live_scrape_cache_size": len(LIVE_SCRAPE_CACHE),
        "live_scrape_cache_ttl_seconds": LIVE_CACHE_TTL_SECONDS,
        "live_scrape_cache_entries": cache_details,
        "last_sync_failed_trh_ids": LAST_SYNC_FAILED_TRH_IDS,
        "api_key_protection_enabled": bool(API_KEY),
        "note": "If cdp_bridge_reachable is false, live-sync/scrape calls will first try to auto-relaunch Chrome "
                "(via GMVN_CHROME_PATH or a common install path) before falling back to FALLBACK_STATIC. "
                "Cache entries older than live_scrape_cache_ttl_seconds are treated as expired and re-scraped."
    }
    dbg(f"/api/debug/status -> cdp_reachable={cdp_reachable}, cache_size={len(LIVE_SCRAPE_CACHE)}")
    return status


@app.get("/favicon.ico")
@app.get("/favicon.png")
def get_favicon():
    fav_path = os.path.join(os.path.dirname(__file__), "favicon.png")
    if os.path.exists(fav_path):
        return FileResponse(fav_path, media_type="image/png")
    raise HTTPException(status_code=404, detail="Favicon not found")


@app.get("/gmvn_master_accommodations.json")
def get_master_json():
    json_path = os.path.join(os.path.dirname(__file__), "gmvn_master_accommodations.json")
    if os.path.exists(json_path):
        return FileResponse(json_path, media_type="application/json")
    return CACHED_DATA


@app.get("/", response_class=HTMLResponse)
def serve_index():
    html_path = os.path.join(os.path.dirname(__file__), "index.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>GMVN API Portal Ready</h1>"


if __name__ == "__main__":
    import uvicorn
    dbg(f"Starting server. DEBUG_ENABLED={DEBUG_ENABLED}. "
        f"Hit GET /api/debug/status anytime to check if live scraping can even work right now.")
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)